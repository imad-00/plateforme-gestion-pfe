import csv
import io
import secrets
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import serializers
from rest_framework.exceptions import PermissionDenied

from apps.academics.models import AcademicYear
from apps.accounts.models import StudentProfile, TeacherProfile, User
from apps.accounts.permissions import get_platform_levels
from apps.audit.models import AdminActionLog
from apps.audit.services import AdminActionLogService
from apps.imports.models import UserImportBatch

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

class UserImportService:
    MAX_ROWS = 1000
    MAX_FILE_SIZE = 5 * 1024 * 1024
    DANGEROUS_PREFIXES = ("=", "+", "-", "@")
    STUDENT_REQUIRED = ["matricule", "email", "first_name", "last_name"]
    STUDENT_COLUMNS = STUDENT_REQUIRED + ["moyenne_generale", "specialite", "academic_year"]
    TEACHER_REQUIRED = ["matricule", "email", "first_name", "last_name"]
    TEACHER_COLUMNS = TEACHER_REQUIRED + ["grade", "departement"]

    @staticmethod
    def _require_admin(actor):
        if not get_platform_levels(actor).intersection({"ADMIN", "SUPER_ADMIN"}):
            raise PermissionDenied("Only platform admins can import users.")

    @staticmethod
    def _is_super_admin(actor):
        return "SUPER_ADMIN" in get_platform_levels(actor)

    @staticmethod
    def _clean(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_header(header):
        return UserImportService._clean(header).lower()

    @staticmethod
    def _row_error(row_number, field, code, message):
        return {"row": row_number, "field": field, "code": code, "message": message}

    @staticmethod
    def _file_error(code, message):
        return {"row": None, "field": None, "code": code, "message": message}

    @staticmethod
    def _validate_import_type(import_type):
        if import_type not in UserImportBatch.ImportType.values:
            raise serializers.ValidationError({"import_type": "Unsupported import type."})

    @staticmethod
    def _validate_upload(uploaded_file):
        if uploaded_file is None:
            raise serializers.ValidationError({"file": "File is required."})
        if getattr(uploaded_file, "size", 0) > UserImportService.MAX_FILE_SIZE:
            raise serializers.ValidationError({"file": "Import file must be 5 MB or smaller."})
        filename = uploaded_file.name or ""
        extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if extension not in {"csv", "xlsx"}:
            raise serializers.ValidationError({"file": "Unsupported file extension. Use CSV or XLSX."})
        return extension

    @staticmethod
    def parse_csv(uploaded_file):
        raw = uploaded_file.read()
        if not raw:
            raise serializers.ValidationError({"file": "Import file is empty."})
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise serializers.ValidationError({"file": "CSV file must be UTF-8 encoded."}) from exc
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise serializers.ValidationError({"file": "CSV file must include a header row."})
        fieldnames = [UserImportService._normalize_header(name) for name in reader.fieldnames]
        rows = []
        for raw_row in reader:
            rows.append(
                {
                    UserImportService._normalize_header(key): UserImportService._clean(value)
                    for key, value in raw_row.items()
                    if key is not None
                }
            )
        return fieldnames, rows

    @staticmethod
    def parse_xlsx(uploaded_file):
        try:
            uploaded_file.seek(0)
            workbook = load_workbook(
                filename=uploaded_file,
                read_only=True,
                data_only=True,
            )
        except InvalidFileException as exc:
            raise serializers.ValidationError({"file": "Invalid XLSX file."}) from exc
        except Exception as exc:
            raise serializers.ValidationError({"file": "Could not read XLSX file."}) from exc

        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration as exc:
            workbook.close()
            raise serializers.ValidationError({"file": "XLSX file is empty."}) from exc

        fieldnames = [UserImportService._normalize_header(value) for value in header]

        if not any(fieldnames):
            workbook.close()
            raise serializers.ValidationError({"file": "XLSX file must include a header row."})

        rows = []

        for excel_row_number, raw_values in enumerate(rows_iter, start=2):
            if raw_values is None or not any(value not in {None, ""} for value in raw_values):
                continue

            row = {}
            for index, field in enumerate(fieldnames):
                if not field:
                    continue
                row[field] = UserImportService._clean(raw_values[index] if index < len(raw_values) else "")

            row["_row_number"] = excel_row_number
            rows.append(row)

        workbook.close()
        return fieldnames, rows

    @staticmethod
    def _parse_file(uploaded_file, extension):
        if extension == "csv":
            return UserImportService.parse_csv(uploaded_file)
        return UserImportService.parse_xlsx(uploaded_file)

    @staticmethod
    def _validate_required_columns(fieldnames, required):
        missing = [column for column in required if column not in fieldnames]
        if missing:
            return [
                UserImportService._file_error(
                    "MISSING_REQUIRED_COLUMN",
                    f"Missing required column: {column}.",
                )
                for column in missing
            ]
        return []

    @staticmethod
    def _validate_common_row(row, row_number, seen_matricules, seen_emails):
        errors = []
        cleaned = {
            "matricule": UserImportService._clean(row.get("matricule")).upper(),
            "email": UserImportService._clean(row.get("email")).lower(),
            "first_name": UserImportService._clean(row.get("first_name")),
            "last_name": UserImportService._clean(row.get("last_name")),
        }
        for field in ["matricule", "email", "first_name", "last_name"]:
            if not cleaned[field]:
                errors.append(UserImportService._row_error(row_number, field, "REQUIRED", "This field is required."))
            if cleaned[field].startswith(UserImportService.DANGEROUS_PREFIXES):
                errors.append(
                    UserImportService._row_error(row_number, field, "SUSPICIOUS_VALUE", "Formula-like values are not allowed.")
                )

        if cleaned["email"]:
            try:
                validate_email(cleaned["email"])
            except DjangoValidationError:
                errors.append(UserImportService._row_error(row_number, "email", "INVALID_EMAIL", "Invalid email address."))

        if cleaned["matricule"] in seen_matricules:
            errors.append(
                UserImportService._row_error(row_number, "matricule", "DUPLICATE_IN_FILE", "Duplicate matricule in file.")
            )
        if cleaned["email"] in seen_emails:
            errors.append(UserImportService._row_error(row_number, "email", "DUPLICATE_IN_FILE", "Duplicate email in file."))
        if cleaned["matricule"]:
            seen_matricules.add(cleaned["matricule"])
        if cleaned["email"]:
            seen_emails.add(cleaned["email"])

        if cleaned["matricule"] and User.objects.filter(matricule__iexact=cleaned["matricule"]).exists():
            errors.append(UserImportService._row_error(row_number, "matricule", "EXISTS", "Matricule already exists."))
        if cleaned["email"] and User.objects.filter(email__iexact=cleaned["email"]).exists():
            errors.append(UserImportService._row_error(row_number, "email", "EXISTS", "Email already exists."))
        return cleaned, errors

    @staticmethod
    def validate_student_row(row, row_number, seen_matricules, seen_emails):
        cleaned, errors = UserImportService._validate_common_row(row, row_number, seen_matricules, seen_emails)
        annual_average = UserImportService._clean(row.get("moyenne_generale"))
        speciality = UserImportService._clean(row.get("specialite"))
        academic_year_label = UserImportService._clean(row.get("academic_year"))

        for field, value in {"specialite": speciality, "academic_year": academic_year_label}.items():
            if value.startswith(UserImportService.DANGEROUS_PREFIXES):
                errors.append(UserImportService._row_error(row_number, field, "SUSPICIOUS_VALUE", "Formula-like values are not allowed."))

        parsed_average = None
        if annual_average:
            try:
                parsed_average = Decimal(annual_average)
                if parsed_average < 0 or parsed_average > 20:
                    raise InvalidOperation
            except (InvalidOperation, ValueError):
                errors.append(
                    UserImportService._row_error(
                        row_number,
                        "moyenne_generale",
                        "INVALID_AVERAGE",
                        "moyenne_generale must be a decimal between 0 and 20.",
                    )
                )

        if academic_year_label:
            academic_year = AcademicYear.objects.filter(year=academic_year_label).first()
            if academic_year is None:
                errors.append(UserImportService._row_error(row_number, "academic_year", "NOT_FOUND", "Academic year not found."))
            elif academic_year.status != AcademicYear.Status.ACTIVE:
                errors.append(
                    UserImportService._row_error(
                        row_number,
                        "academic_year",
                        "NOT_ACTIVE",
                        "Student import academic year must be ACTIVE.",
                    )
                )
        else:
            academic_year = AcademicYear.objects.filter(status=AcademicYear.Status.ACTIVE).first()
            if academic_year is None:
                errors.append(
                    UserImportService._row_error(row_number, "academic_year", "NO_ACTIVE_YEAR", "No active academic year is configured.")
                )

        normalized = {
            **cleaned,
            "business_identity": User.BusinessIdentity.STUDENT,
            "annual_average": str(parsed_average) if parsed_average is not None else None,
            "speciality": speciality,
            "academic_year_id": academic_year.id if "academic_year" in locals() and academic_year else None,
            "academic_year": academic_year.year if "academic_year" in locals() and academic_year else "",
        }
        return normalized, errors

    @staticmethod
    def validate_teacher_row(row, row_number, seen_matricules, seen_emails):
        cleaned, errors = UserImportService._validate_common_row(row, row_number, seen_matricules, seen_emails)
        grade = UserImportService._clean(row.get("grade"))
        department = UserImportService._clean(row.get("departement"))
        for field, value in {"grade": grade, "departement": department}.items():
            if value.startswith(UserImportService.DANGEROUS_PREFIXES):
                errors.append(UserImportService._row_error(row_number, field, "SUSPICIOUS_VALUE", "Formula-like values are not allowed."))
        return {
            **cleaned,
            "business_identity": User.BusinessIdentity.TEACHER,
            "grade": grade,
            "department": department,
        }, errors

    @staticmethod
    def preview_user_import(actor, uploaded_file, import_type, request=None):
        UserImportService._require_admin(actor)
        UserImportService._validate_import_type(import_type)
        extension = UserImportService._validate_upload(uploaded_file)
        fieldnames, rows = UserImportService._parse_file(uploaded_file, extension)
        if not rows:
            raise serializers.ValidationError({"file": "Import file contains no data rows."})
        if len(rows) > UserImportService.MAX_ROWS:
            raise serializers.ValidationError({"file": f"Import file cannot exceed {UserImportService.MAX_ROWS} rows."})

        required = UserImportService.STUDENT_REQUIRED if import_type == UserImportBatch.ImportType.STUDENTS else UserImportService.TEACHER_REQUIRED
        errors = UserImportService._validate_required_columns(fieldnames, required)
        normalized_rows = []
        seen_matricules = set()
        seen_emails = set()

        for index, row in enumerate(rows, start=2):
            row_number = row.get("_row_number", index)

            if import_type == UserImportBatch.ImportType.STUDENTS:
                normalized, row_errors = UserImportService.validate_student_row(row, row_number, seen_matricules, seen_emails)
            else:
                normalized, row_errors = UserImportService.validate_teacher_row(row, row_number, seen_matricules, seen_emails)

            normalized["row_number"] = row_number
            normalized["is_valid"] = not row_errors
            normalized_rows.append(normalized)
            errors.extend(row_errors)
            
        valid_count = sum(1 for row in normalized_rows if row["is_valid"])
        batch = UserImportBatch.objects.create(
            import_type=import_type,
            uploaded_by=actor,
            original_filename=uploaded_file.name or "",
            total_rows=len(rows),
            valid_rows=valid_count,
            invalid_rows=len(rows) - valid_count,
            errors=errors,
            warnings=[],
            normalized_rows=normalized_rows,
        )
        AdminActionLogService.log(
            actor,
            AdminActionLog.ActionType.USER_IMPORT_PREVIEWED,
            target=batch,
            metadata={
                "import_type": import_type,
                "total_rows": batch.total_rows,
                "valid_rows": batch.valid_rows,
                "invalid_rows": batch.invalid_rows,
                "original_filename": batch.original_filename,
            },
            request=request,
        )
        return batch

    @staticmethod
    def _assert_can_confirm(actor, batch):
        if batch.uploaded_by_id != actor.id and not UserImportService._is_super_admin(actor):
            raise PermissionDenied("Only the upload owner or a super admin can confirm this import.")
        if batch.status != UserImportBatch.Status.PREVIEWED:
            raise serializers.ValidationError({"batch_id": "Only previewed import batches can be confirmed."})

    @staticmethod
    def _generate_password():
        return secrets.token_urlsafe(24)

    @staticmethod
    def create_student_from_row(row):
        user = User.objects.create_user(
            matricule=row["matricule"],
            email=row["email"],
            password=UserImportService._generate_password(),
            first_name=row["first_name"],
            last_name=row["last_name"],
            business_identity=User.BusinessIdentity.STUDENT,
            account_status=User.AccountStatus.ACTIVE,
            must_reset_password=True,
        )
        StudentProfile.objects.update_or_create(
            user=user,
            defaults={
                "academic_year_id": row["academic_year_id"],
                "annual_average": Decimal(row["annual_average"]) if row.get("annual_average") else None,
                "speciality": row.get("speciality", ""),
            },
        )
        from apps.teams.services import TeamService

        TeamService.create_solo_team_for_student(
            user,
            academic_year=AcademicYear.objects.get(pk=row["academic_year_id"]),
        )
        return user

    @staticmethod
    def create_teacher_from_row(row):
        user = User.objects.create_user(
            matricule=row["matricule"],
            email=row["email"],
            password=UserImportService._generate_password(),
            first_name=row["first_name"],
            last_name=row["last_name"],
            business_identity=User.BusinessIdentity.TEACHER,
            account_status=User.AccountStatus.ACTIVE,
            must_reset_password=True,
        )
        TeacherProfile.objects.update_or_create(
            user=user,
            defaults={
                "grade": row.get("grade", ""),
                "department": row.get("department", ""),
            },
        )
        return user

    @staticmethod
    @transaction.atomic
    def confirm_user_import(actor, batch, confirm=False, allow_partial=False, request=None):
        UserImportService._require_admin(actor)
        UserImportService._assert_can_confirm(actor, batch)
        if confirm is not True:
            raise serializers.ValidationError({"confirm": "confirm must be true."})
        if batch.invalid_rows and not allow_partial:
            raise serializers.ValidationError({"batch_id": "Import batch has invalid rows. Use allow_partial=true to import valid rows only."})

        valid_rows = [row for row in batch.normalized_rows if row.get("is_valid")]
        created_users = []
        for row in valid_rows:
            if batch.import_type == UserImportBatch.ImportType.STUDENTS:
                user = UserImportService.create_student_from_row(row)
            else:
                user = UserImportService.create_teacher_from_row(row)
            created_users.append(user)
            AdminActionLogService.log(
                actor,
                AdminActionLog.ActionType.USER_CREATED_BY_IMPORT,
                target=user,
                metadata={
                    "batch_id": batch.id,
                    "import_type": batch.import_type,
                    "row_number": row.get("row_number"),
                    "business_identity": user.business_identity,
                },
                request=request,
            )

        batch.status = UserImportBatch.Status.COMPLETED
        batch.created_count = len(created_users)
        batch.skipped_count = batch.invalid_rows if allow_partial else 0
        batch.completed_at = timezone.now()
        batch.save(update_fields=["status", "created_count", "skipped_count", "completed_at", "updated_at"])
        AdminActionLogService.log(
            actor,
            AdminActionLog.ActionType.USER_IMPORT_COMPLETED,
            target=batch,
            metadata={
                "import_type": batch.import_type,
                "created_count": batch.created_count,
                "skipped_count": batch.skipped_count,
                "error_count": batch.invalid_rows,
            },
            request=request,
        )
        return {
            "batch": batch,
            "created_count": batch.created_count,
            "skipped_count": batch.skipped_count,
            "error_count": batch.invalid_rows,
            "created_users": [
                {
                    "id": user.id,
                    "matricule": user.matricule,
                    "email": user.email,
                    "business_identity": user.business_identity,
                    "must_reset_password": user.must_reset_password,
                }
                for user in created_users
            ],
        }

    @staticmethod
    def generate_template(import_type):
        UserImportService._validate_import_type(import_type)
        columns = (
            UserImportService.STUDENT_COLUMNS
            if import_type == UserImportBatch.ImportType.STUDENTS
            else UserImportService.TEACHER_COLUMNS
        )
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
        filename = "students_import_template.csv" if import_type == UserImportBatch.ImportType.STUDENTS else "teachers_import_template.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
